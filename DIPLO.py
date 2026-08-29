#!/usr/bin/env python3

from datetime import datetime

# ============================================================
# Interactive Multi-Platform Search + LLM Translator
# + Location Inference + Interactive Map
# ============================================================
#
# What this does:
# 1. Prompts for any search terms and optional query-translation languages.
# 2. Searches public posts on X, Bluesky, Mastodon, or multiple platforms.
# 3. Normalizes each platform's structured API response.
# 4. Translates posts using the selected LLM provider.
# 5. Infers broad public locations using the selected LLM provider.
# 6. Geocodes inferred locations with Nominatim / OpenStreetMap.
# 7. Exports clean CSV and formatted Excel files.
# 8. Creates an interactive HTML map with popups.
#
# This does NOT bypass login walls, CAPTCHAs, private accounts,
# anti-bot systems, access controls, or rate limits. 
#
# HOW TO RUN (macOS / Linux Terminal)
# -----------------------------------
# First-time setup, from your project directory:
#
#     python3 -m venv .venv
#     source .venv/bin/activate
#     python3 -m pip install --upgrade pip
#     python3 DIPLO.py
#
# The script installs missing Python packages into the active virtual
# environment. On later runs, from your project directory:
#
#     source .venv/bin/activate
#     python3 DIPLO.py
#
# At startup, choose one or more social-data sources. X requires an App-Only
# Bearer Token. Bluesky can use public search, or a handle and Bluesky app
# password when the public API is blocked. Mastodon accepts an optional user
# access token. Then choose an LLM provider for translation and enrichment.
# All credentials use hidden input, remain in memory only for the current run,
# and are not written to project files or outputs.
#
# To leave the virtual environment afterward:
#
#     deactivate
#
# ============================================================


# ============================================================
# USER SETTINGS
# ============================================================

# Use api mode to choose X, Bluesky, Mastodon, or a combination.
# Use saved_html only for legacy Nitter HTML files already saved on disk.
MODE = "api"
# MODE = "saved_html"

X_API_BASE_URL = "https://api.x.com/2"
X_BEARER_TOKEN = ""
BLUESKY_API_BASE_URL = "https://public.api.bsky.app"
BLUESKY_PDS_URL = "https://bsky.social"
BLUESKY_SERVICE_PROXY = "did:web:api.bsky.app#bsky_appview"
BLUESKY_IDENTIFIER = ""
BLUESKY_APP_PASSWORD = ""
BLUESKY_ACCESS_JWT = ""
MASTODON_INSTANCE_URL = "https://mastodon.social"
MASTODON_ACCESS_TOKEN = ""

# For saved_html mode, save Nitter pages from your browser and list them here.
SAVED_HTML_FILES = [
    # "nitter_results.html",
]

LLM_MODEL = "gpt-5.6-luna"
LLM_PROVIDER = "openai"
LLM_BASE_URL = ""
LLM_API_KEY = ""
TARGET_LANGUAGE = "English"

# Start broad. Once results work, add dates.
SINCE_DATE = None
UNTIL_DATE = None

# Example:
# SINCE_DATE = "2026-01-01"
# UNTIL_DATE = "2026-06-29"

MAX_POSTS_PER_QUERY = 10
MAX_PAGES_PER_QUERY = 1

INCLUDE_RETWEETS = False
TRANSLATE_POSTS = True

REQUEST_DELAY_SECONDS = 3.0
TRANSLATION_DELAY_SECONDS = 0.5

# One timestamp is created when the script starts and shared by every primary
# output from that run. Example: social_search_posts_20260829_143012.csv
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = f"social_search_posts_{RUN_TIMESTAMP}.csv"

# Leave empty to search keywords.
# Add handles for cleaner results from known accounts.
SEARCH_HANDLES = [
    # "@ConfuciusINST",
    # "@SomeUniversityCI",
]

DEFAULT_SEARCH_TERMS = [
    "Democracy",
]

# The original terms are always searched. The user can add translations in any
# of these languages at runtime by entering the corresponding number(s).
SEARCH_LANGUAGE_OPTIONS = {
    "1": "English",
    "2": "Simplified Chinese",
    "3": "Traditional Chinese",
    "4": "Spanish",
    "5": "French",
    "6": "Portuguese",
    "7": "Italian",
    "8": "German",
    "9": "Russian",
    "10": "Arabic",
    "11": "Japanese",
    "12": "Korean",
    "13": "Hindi",
    "14": "Turkish",
    "15": "Indonesian",
}


# ============================================================
# MAP / LOCATION SETTINGS
# ============================================================

CREATE_MAP = True
INFER_LOCATIONS = True

# The geocoding cache intentionally keeps a stable name so later runs can reuse
# it. It is a support cache, not a run output.
MAP_FILE = f"social_search_map_{RUN_TIMESTAMP}.html"
GEOCODE_CACHE_FILE = "geocode_cache.json"

# Map tile settings.
# The old Folium default, tiles="OpenStreetMap", can request tiles from
# tile.openstreetmap.org and may produce browser errors such as:
# 503r: Access Blocked / osm.wiki/blocked.
# To avoid that, this script does NOT use the direct OpenStreetMap tile server.
# It uses CARTO's public basemap CDN instead, with proper attribution.
MAP_BASE_TILE_NAME = "CartoDB Positron"
MAP_BASE_TILE_URL = "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png"
MAP_BASE_TILE_ATTRIBUTION = (
    '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> '
    'contributors &copy; <a href="https://carto.com/attributions">CARTO</a>'
)

# Optional fallback/alternate basemaps. These also avoid the direct OSM tile server.
ADD_ALTERNATE_BASEMAPS = True

# Inferred locations are broad, public, city/region/country-level locations.
# They are not verified geotags.
MIN_LOCATION_CONFIDENCE = 0.2

# Nominatim requires a real identifying User-Agent.
NOMINATIM_USER_AGENT = "interactive-nitter-search-map-research/1.0"

# Nominatim public service requires no more than 1 request per second.
GEOCODE_DELAY_SECONDS = 1.2


# ============================================================
# INSTALL PACKAGES
# ============================================================

import sys
import subprocess
import importlib.util
import os
import time
import csv
import re
import json
import html as html_lib
from getpass import getpass
from dataclasses import asdict, dataclass
from typing import List, Optional, Dict, Any
from urllib.parse import urlencode, urljoin, urlparse



def install_if_missing(import_name, pip_name=None):
    if pip_name is None:
        pip_name = import_name

    if importlib.util.find_spec(import_name) is None:
        print(f"Installing {pip_name}...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q", pip_name]
        )
    else:
        print(f"{pip_name} already installed.")


install_if_missing("openai")
install_if_missing("pandas")
install_if_missing("bs4", "beautifulsoup4")
install_if_missing("requests")
install_if_missing("certifi")
install_if_missing("langdetect")
install_if_missing("tqdm")
install_if_missing("dateutil", "python-dateutil")
install_if_missing("openpyxl")
install_if_missing("folium")
install_if_missing("geopy")
install_if_missing("python-dotenv")


# ============================================================
# FIX SSL CERTIFICATE ISSUES
# ============================================================

print("Upgrading certificate-related packages...")

subprocess.check_call([
    sys.executable, "-m", "pip", "install", "-q", "--upgrade",
    "certifi", "requests", "urllib3"
])

import certifi

os.environ["SSL_CERT_FILE"] = certifi.where()
os.environ["REQUESTS_CA_BUNDLE"] = certifi.where()
os.environ["CURL_CA_BUNDLE"] = certifi.where()

print("Using certifi CA bundle:")
print(certifi.where())


# ============================================================
# IMPORT PACKAGES
# ============================================================

import pandas as pd
import requests
from bs4 import BeautifulSoup
from langdetect import detect, LangDetectException
from openai import OpenAI
from dateutil import parser as dateparser
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import folium
from folium.plugins import MarkerCluster
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderUnavailable

try:
    from tqdm import tqdm
except Exception:
    tqdm = None


# ============================================================
# LLM PROVIDER, MODEL, AND API KEY
# ============================================================

ARC_BASE_URL = "https://llm-api.arc.vt.edu/api/v1"


def prompt_required_secret(label: str) -> str:
    """Read a required secret without echoing or persisting it."""
    while True:
        value = getpass(label).strip()
        if value:
            return value
        print("An API key is required.")


def prompt_llm_settings():
    """Choose OpenAI or Virginia Tech ARC, a model, and a runtime API key."""
    print("\nLLM provider:")
    print("  1. OpenAI API (default)")
    print("  2. Virginia Tech ARC LLM API")

    while True:
        provider_choice = input("Choose provider [1]: ").strip() or "1"
        if provider_choice in {"1", "2"}:
            break
        print("Please enter 1 or 2.")

    if provider_choice == "1":
        models = {
            "1": "gpt-5.6-luna",
            "2": "gpt-5.6-terra",
            "3": "gpt-5.6-sol",
        }
        print("\nOpenAI model:")
        print("  1. gpt-5.6-luna — cost-sensitive/high-volume (default)")
        print("  2. gpt-5.6-terra — balance of capability and cost")
        print("  3. gpt-5.6-sol — flagship capability")
        print("  4. Enter another OpenAI model ID")

        while True:
            model_choice = input("Choose OpenAI model [1]: ").strip() or "1"
            if model_choice in models:
                model = models[model_choice]
                break
            if model_choice == "4":
                model = input("OpenAI model ID: ").strip()
                if model:
                    break
            print("Please enter a number from 1 to 4.")

        print("Enter an OpenAI Platform API key; input is hidden and not saved.")
        api_key = prompt_required_secret("OpenAI API key: ")
        return "openai", model, "", api_key

    models = {
        "1": "gpt-oss-120b",
        "2": "DeepSeek-V4-Flash",
        "3": "GLM-5.2",
        "4": "Kimi-K3",
    }
    print("\nVirginia Tech ARC model:")
    print("  1. gpt-oss-120b (default)")
    print("  2. DeepSeek-V4-Flash")
    print("  3. GLM-5.2")
    print("  4. Kimi-K3")
    print("  5. Enter another ARC model ID")

    while True:
        model_choice = input("Choose ARC model [1]: ").strip() or "1"
        if model_choice in models:
            model = models[model_choice]
            break
        if model_choice == "5":
            model = input("ARC model ID: ").strip()
            if model:
                break
        print("Please enter a number from 1 to 5.")

    print(
        "Create a personal key at https://llm.arc.vt.edu under "
        "User profile > Settings > Account > API keys."
    )
    print("Input is hidden and the key is not saved.")
    api_key = prompt_required_secret("ARC personal API key: ")
    return "arc", model, ARC_BASE_URL, api_key


def create_llm_client() -> OpenAI:
    """Create a client for the provider selected at startup."""
    options = {"api_key": LLM_API_KEY}
    if LLM_BASE_URL:
        options["base_url"] = LLM_BASE_URL
    return OpenAI(**options)


def generate_llm_text(client: OpenAI, model: str, prompt: str) -> str:
    """Generate text through OpenAI Responses or ARC Chat Completions."""
    if LLM_PROVIDER == "arc":
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=8_000,
        )
        return (response.choices[0].message.content or "").strip()

    response = client.responses.create(model=model, input=prompt)
    return response.output_text.strip()


def ensure_llm_configured():
    """Prompt once, including when only search-term translation needs an LLM."""
    global LLM_PROVIDER, LLM_MODEL, LLM_BASE_URL, LLM_API_KEY
    if LLM_API_KEY:
        return
    LLM_PROVIDER, LLM_MODEL, LLM_BASE_URL, LLM_API_KEY = prompt_llm_settings()
    print(f"\nUsing {LLM_PROVIDER.upper()} model {LLM_MODEL}.")
    print("API key accepted for this run (not saved).")


# ============================================================
# DATA STRUCTURE
# ============================================================

@dataclass
class PostRecord:
    platform: str
    query: str
    source_mode: str
    source_host: str
    source_url: str
    scraped_from: str
    post_url: str
    x_url: str
    tweet_id: str
    date_raw: str
    date_iso: str
    username: str
    display_name: str
    detected_language: str
    original_text: str
    translated_en: str
    raw_stats: str
    is_retweet: bool
    inferred_location: str
    location_confidence: float
    location_source: str
    location_reason: str
    latitude: str
    longitude: str
    geocode_display_name: str


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_handle(handle: str) -> str:
    handle = handle.strip()
    if handle.startswith("@"):
        handle = handle[1:]
    return handle


def normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def clean_cell(value):
    """
    Make social media text safer and cleaner for CSV/Excel output.
    Removes internal line breaks, normalizes spacing, and prevents
    spreadsheet formula injection.
    """

    if value is None:
        return ""

    value = str(value)

    value = value.replace("\r", " ")
    value = value.replace("\n", " ")
    value = value.replace("\t", " ")

    value = normalize_whitespace(value)

    # Prevent spreadsheet formula injection
    if value.startswith(("=", "+", "-", "@")):
        value = "'" + value

    return value


def detect_language_safe(text: str) -> str:
    try:
        return detect(text)
    except LangDetectException:
        return "unknown"
    except Exception:
        return "unknown"


def parse_nitter_date(date_text: str) -> str:
    if not date_text:
        return ""

    cleaned = date_text.replace("·", " ").strip()

    try:
        dt = dateparser.parse(cleaned, fuzzy=True)
        if dt:
            return dt.isoformat()
    except Exception:
        return ""

    return ""


def build_queries(
    terms: List[str],
    handles: Optional[List[str]] = None,
) -> List[str]:
    if handles:
        queries = []

        for handle in handles:
            h = clean_handle(handle)
            queries.append(f"from:{h}")

        return queries

    return [term.strip() for term in terms if term.strip()]


def prompt_x_bearer_token() -> str:
    """Read the X App-Only Bearer Token without echoing or persisting it."""
    print("\nX API authentication")
    print(
        "Use the App-Only Bearer Token from the X Developer Console under "
        "your app's Keys & Tokens page."
    )
    print("The token is hidden while entered and is not saved by this script.")
    return prompt_required_secret("X Bearer Token: ")


def prompt_data_sources() -> List[str]:
    """Let the user select one or more social-data APIs."""
    print("\nSocial-data sources:")
    print("  1. X API")
    print("  2. Bluesky (public API or app-password login)")
    print("  3. Mastodon API")
    print("  4. All three")
    print("You may also enter a comma-separated combination, such as 1,2.")

    while True:
        raw = input("Choose data source(s) [1]: ").strip() or "1"
        if raw == "4":
            return ["x", "bluesky", "mastodon"]

        mapping = {"1": "x", "2": "bluesky", "3": "mastodon"}
        choices = [part.strip() for part in raw.split(",") if part.strip()]
        if choices and all(choice in mapping for choice in choices):
            selected = []
            for choice in choices:
                source = mapping[choice]
                if source not in selected:
                    selected.append(source)
            return selected
        print("Enter 1, 2, 3, 4, or a combination such as 1,2.")


def prompt_bluesky_settings() -> tuple[str, str]:
    """Optionally collect a Bluesky handle and app password for PDS proxying."""
    print("\nBluesky search setup")
    print("Public Bluesky search normally needs no login, but some networks block")
    print("the public *.bsky.app API hosts. An app-password login uses bsky.social")
    print("as a reachable authenticated proxy when that happens.")
    print("Create an app password in Bluesky Settings > Privacy and Security >")
    print("App Passwords. Do not enter your main Bluesky password here.")
    identifier = input(
        "Bluesky handle or account email (press Enter for public-only): "
    ).strip()
    if not identifier:
        return "", ""
    app_password = prompt_required_secret("Bluesky app password (input hidden): ")
    return identifier, app_password


def create_bluesky_access_token(
    session: requests.Session,
    identifier: str,
    app_password: str,
) -> str:
    """Create an in-memory Bluesky session and return its short-lived JWT."""
    endpoint = f"{BLUESKY_PDS_URL.rstrip('/')}/xrpc/com.atproto.server.createSession"
    response = session.post(
        endpoint,
        json={"identifier": identifier, "password": app_password},
        headers={"Accept": "application/json"},
        timeout=60,
    )
    try:
        payload = response.json()
    except ValueError as exc:
        raise RuntimeError(
            f"Bluesky login returned a non-JSON response ({response.status_code})."
        ) from exc
    if response.status_code in {400, 401}:
        message = payload.get("message") or payload.get("error") or "login rejected"
        raise RuntimeError(f"Bluesky rejected the handle or app password: {message}")
    try:
        response.raise_for_status()
    except requests.RequestException as exc:
        message = payload.get("message") or payload.get("error") or "unknown error"
        raise RuntimeError(
            f"Bluesky login failed ({response.status_code}): {message}"
        ) from exc
    token = str(payload.get("accessJwt", "")).strip()
    if not token:
        raise RuntimeError("Bluesky login succeeded but returned no access token.")
    return token


def prompt_mastodon_settings() -> tuple[str, str]:
    """Choose a Mastodon server and optionally read a user access token."""
    print("\nMastodon search setup")
    print(
        "Mastodon searches one server's known/indexed posts, not the entire "
        "Fediverse. Full-text status search normally works best with a user token."
    )
    instance = input(
        f"Mastodon server URL [{MASTODON_INSTANCE_URL}]: "
    ).strip() or MASTODON_INSTANCE_URL
    if not instance.startswith(("https://", "http://")):
        instance = "https://" + instance
    instance = instance.rstrip("/")

    token = getpass(
        "Mastodon user access token (input hidden; press Enter for none): "
    ).strip()
    if not token:
        print(
            "Continuing without a Mastodon token. The server may return no "
            "full-text status results or may restrict pagination."
        )
    return instance, token


def x_api_time(value: Optional[str]) -> Optional[str]:
    """Convert a YYYY-MM-DD setting to the RFC 3339 form expected by X."""
    if not value:
        return None
    value = value.strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return value + "T00:00:00Z"
    return value


def build_x_query(query: str, include_retweets: bool) -> str:
    """Add the retweet filter unless the user already supplied one."""
    query = normalize_whitespace(query)
    if not include_retweets and not re.search(r"(?:^|\s)-?is:retweet(?:\s|$)", query):
        query = f"{query} -is:retweet"
    return query


def fetch_x_api_page(
    session: requests.Session,
    bearer_token: str,
    query: str,
    max_results: int,
    since: Optional[str] = None,
    until: Optional[str] = None,
    next_token: Optional[str] = None,
) -> tuple[List[Dict[str, Any]], Optional[str], str]:
    """Fetch and normalize one page from X API v2 recent search."""
    endpoint = f"{X_API_BASE_URL.rstrip('/')}/tweets/search/recent"
    params: Dict[str, Any] = {
        "query": query,
        "max_results": max(10, min(100, max_results)),
        "tweet.fields": (
            "id,text,author_id,created_at,lang,public_metrics,referenced_tweets"
        ),
        "expansions": "author_id",
        "user.fields": "id,name,username,location",
    }
    if since:
        params["start_time"] = x_api_time(since)
    if until:
        params["end_time"] = x_api_time(until)
    if next_token:
        params["next_token"] = next_token

    response = session.get(
        endpoint,
        params=params,
        headers={"Authorization": f"Bearer {bearer_token}"},
        timeout=60,
    )

    if response.status_code == 401:
        raise RuntimeError("X rejected the Bearer Token (401 Unauthorized).")
    if response.status_code == 402:
        raise RuntimeError(
            "X reported that API credits or billing are required (402)."
        )
    if response.status_code == 403:
        raise RuntimeError(
            "X denied access to recent search for this app (403 Forbidden)."
        )
    if response.status_code == 429:
        raise RuntimeError("X API rate limit reached (429). Try again later.")

    try:
        response.raise_for_status()
        payload = response.json()
    except requests.RequestException as exc:
        detail = response.text[:500] if response is not None else ""
        raise RuntimeError(f"X API request failed: {exc}. {detail}") from exc
    except ValueError as exc:
        raise RuntimeError("X API returned a non-JSON response.") from exc

    errors = payload.get("errors") or []
    if errors and not payload.get("data"):
        raise RuntimeError(f"X API returned errors: {json.dumps(errors)}")

    users = {
        str(user.get("id", "")): user
        for user in payload.get("includes", {}).get("users", [])
    }
    posts: List[Dict[str, Any]] = []

    for item in payload.get("data", []) or []:
        author = users.get(str(item.get("author_id", "")), {})
        username = str(author.get("username", "")).strip()
        tweet_id = str(item.get("id", "")).strip()
        referenced = item.get("referenced_tweets") or []
        is_retweet = any(ref.get("type") == "retweeted" for ref in referenced)
        created_at = str(item.get("created_at", "")).strip()
        public_metrics = item.get("public_metrics") or {}
        x_url = (
            f"https://x.com/{username}/status/{tweet_id}"
            if username
            else f"https://x.com/i/web/status/{tweet_id}"
        )

        posts.append(
            {
                "platform": "x",
                "query": query,
                "scraped_from": response.url,
                "source_url": response.url,
                "post_url": x_url,
                "x_url": x_url,
                "tweet_id": tweet_id,
                "date_raw": created_at,
                "date_iso": created_at,
                "username": username,
                "display_name": str(author.get("name", "")).strip(),
                "original_text": str(item.get("text", "")),
                "raw_stats": json.dumps(public_metrics, sort_keys=True),
                "is_retweet": is_retweet,
            }
        )

    pagination_token = payload.get("meta", {}).get("next_token")
    return posts, pagination_token, response.url


def fetch_bluesky_page(
    session: requests.Session,
    query: str,
    max_results: int,
    since: Optional[str] = None,
    until: Optional[str] = None,
    cursor: Optional[str] = None,
    access_jwt: str = "",
) -> tuple[List[Dict[str, Any]], Optional[str], str]:
    """Fetch one Bluesky page publicly or through an authenticated PDS proxy."""
    if access_jwt:
        endpoint = f"{BLUESKY_PDS_URL.rstrip('/')}/xrpc/app.bsky.feed.searchPosts"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_jwt}",
            "atproto-proxy": BLUESKY_SERVICE_PROXY,
        }
    else:
        endpoint = (
            f"{BLUESKY_API_BASE_URL.rstrip('/')}/xrpc/app.bsky.feed.searchPosts"
        )
        headers = {"Accept": "application/json"}
    params: Dict[str, Any] = {
        "q": query,
        "sort": "latest",
        "limit": max(1, min(100, max_results)),
    }
    if since:
        params["since"] = since
    if until:
        params["until"] = until
    if cursor:
        params["cursor"] = cursor

    response = session.get(endpoint, params=params, headers=headers, timeout=60)
    try:
        response.raise_for_status()
        payload = response.json()
    except requests.RequestException as exc:
        content_type = response.headers.get("Content-Type", "").lower()
        if response.status_code == 403 and "html" in content_type:
            raise RuntimeError(
                "the public Bluesky API was blocked by a network filter (HTML "
                "403). Run the script again and enter your Bluesky handle plus "
                "an app password, or use a network/VPN that permits *.bsky.app."
            ) from exc
        detail = normalize_whitespace(response.text[:300])
        raise RuntimeError(
            f"Bluesky API request failed ({response.status_code}): "
            f"{detail}"
        ) from exc
    except ValueError as exc:
        raise RuntimeError("Bluesky returned a non-JSON response.") from exc

    posts: List[Dict[str, Any]] = []
    for item in payload.get("posts", []) or []:
        author = item.get("author") or {}
        record = item.get("record") or {}
        uri = str(item.get("uri", ""))
        rkey = uri.rsplit("/", 1)[-1] if "/" in uri else ""
        handle = str(author.get("handle", "")).strip()
        post_url = (
            f"https://bsky.app/profile/{handle}/post/{rkey}"
            if handle and rkey
            else ""
        )
        metrics = {
            "reply_count": item.get("replyCount", 0),
            "repost_count": item.get("repostCount", 0),
            "like_count": item.get("likeCount", 0),
            "quote_count": item.get("quoteCount", 0),
        }
        created_at = str(record.get("createdAt", item.get("indexedAt", "")))

        posts.append(
            {
                "platform": "bluesky",
                "query": query,
                "scraped_from": response.url,
                "source_url": response.url,
                "post_url": post_url,
                "x_url": "",
                "tweet_id": rkey or uri,
                "date_raw": created_at,
                "date_iso": created_at,
                "username": handle,
                "display_name": str(author.get("displayName", "")).strip(),
                "original_text": str(record.get("text", "")),
                "raw_stats": json.dumps(metrics, sort_keys=True),
                "is_retweet": False,
            }
        )

    return posts, payload.get("cursor"), response.url


def mastodon_plain_text(content: str) -> str:
    """Convert a Mastodon status's HTML body into readable plain text."""
    return normalize_whitespace(BeautifulSoup(content or "", "html.parser").get_text(" "))


def date_is_in_range(
    created_at: str,
    since: Optional[str],
    until: Optional[str],
) -> bool:
    """Apply optional ISO-like date bounds to APIs without date parameters."""
    if not created_at:
        return True
    try:
        created = dateparser.parse(created_at)
        lower = dateparser.parse(since) if since else None
        upper = dateparser.parse(until) if until else None
        if created is None:
            return True
        if lower and created.date() < lower.date():
            return False
        if upper and created.date() >= upper.date():
            return False
        return True
    except Exception:
        return True


def fetch_mastodon_page(
    session: requests.Session,
    instance_url: str,
    access_token: str,
    query: str,
    max_results: int,
    since: Optional[str] = None,
    until: Optional[str] = None,
    offset: int = 0,
) -> tuple[List[Dict[str, Any]], Optional[int], str]:
    """Fetch one instance-scoped Mastodon full-text status search page."""
    endpoint = f"{instance_url.rstrip('/')}/api/v2/search"
    params: Dict[str, Any] = {
        "q": query,
        "type": "statuses",
        "limit": max(1, min(40, max_results)),
    }
    if access_token and offset:
        params["offset"] = offset
    headers = {"Authorization": f"Bearer {access_token}"} if access_token else {}

    response = session.get(endpoint, params=params, headers=headers, timeout=60)
    try:
        response.raise_for_status()
        payload = response.json()
    except requests.RequestException as exc:
        raise RuntimeError(
            f"Mastodon API request failed ({response.status_code}): "
            f"{response.text[:500]}"
        ) from exc
    except ValueError as exc:
        raise RuntimeError("Mastodon returned a non-JSON response.") from exc

    statuses = payload.get("statuses", []) or []
    posts: List[Dict[str, Any]] = []
    for status in statuses:
        is_reblog = bool(status.get("reblog"))
        content_status = status.get("reblog") or status
        account = content_status.get("account") or status.get("account") or {}
        created_at = str(content_status.get("created_at", status.get("created_at", "")))
        if not date_is_in_range(created_at, since, until):
            continue

        metrics = {
            "reply_count": content_status.get("replies_count", 0),
            "reblog_count": content_status.get("reblogs_count", 0),
            "favourite_count": content_status.get("favourites_count", 0),
        }
        post_url = str(content_status.get("url", status.get("url", "")))
        posts.append(
            {
                "platform": "mastodon",
                "query": query,
                "scraped_from": response.url,
                "source_url": response.url,
                "post_url": post_url,
                "x_url": "",
                "tweet_id": str(content_status.get("id", status.get("id", ""))),
                "date_raw": created_at,
                "date_iso": created_at,
                "username": str(account.get("acct", account.get("username", ""))),
                "display_name": mastodon_plain_text(
                    str(account.get("display_name", ""))
                ),
                "original_text": mastodon_plain_text(
                    str(content_status.get("content", ""))
                ),
                "raw_stats": json.dumps(metrics, sort_keys=True),
                "is_retweet": is_reblog,
            }
        )

    next_offset = None
    if access_token and len(statuses) == params["limit"]:
        next_offset = offset + len(statuses)
    return posts, next_offset, response.url


def make_nitter_search_url(
    base_url: str,
    query: str,
    since: Optional[str] = None,
    until: Optional[str] = None,
) -> str:
    base_url = base_url.rstrip("/")

    q = query.strip()

    q = re.sub(r"\bsince:\d{4}-\d{2}-\d{2}\b", "", q)
    q = re.sub(r"\buntil:\d{4}-\d{2}-\d{2}\b", "", q)
    q = q.replace("-filter:retweets", "")
    q = normalize_whitespace(q)

    if q.startswith('"') and q.endswith('"'):
        q = q[1:-1]

    params = {
        "f": "tweets",
        "q": q,
    }

    if since:
        params["since"] = since

    if until:
        params["until"] = until

    return f"{base_url}/search?{urlencode(params)}"


def make_x_url_from_nitter_url(nitter_url: str) -> str:
    try:
        parsed = urlparse(nitter_url)
        return "https://x.com" + parsed.path
    except Exception:
        return ""


def extract_tweet_id_from_url(url: str) -> str:
    match = re.search(r"/status/(\d+)", url)
    if match:
        return match.group(1)
    return ""


def get_text_or_empty(element) -> str:
    if element is None:
        return ""
    return element.get_text(" ", strip=True)


def prompt_for_search_terms(default_terms: List[str]) -> List[str]:
    """Collect arbitrary search terms, one per line, from the user."""

    print("\n" + "=" * 70)
    print("SEARCH SETUP")
    print("=" * 70)
    print("Enter any topics, names, phrases, or hashtags to search for.")
    print("Enter one term per line, then press Enter on a blank line to finish.")
    print(f"Press Enter immediately to use the default: {', '.join(default_terms)}")

    terms = []

    while True:
        try:
            value = input("Search term: ").strip()
        except EOFError:
            print("No interactive input available; using the default search terms.")
            return list(default_terms)

        if not value:
            break

        if value not in terms:
            terms.append(value)

    return terms or list(default_terms)


def prompt_for_search_languages() -> List[str]:
    """Let the user select languages for translated search-query variants."""

    print("\nThe original search terms will always be used.")
    print("Optionally translate them into additional search languages:")

    for number, language in SEARCH_LANGUAGE_OPTIONS.items():
        print(f"  {number:>2}. {language}")

    print("Enter numbers separated by commas (for example: 2,4,6).")
    print("Press Enter to search only the original terms.")

    try:
        raw_choices = input("Language choices: ").strip()
    except EOFError:
        return []

    if not raw_choices:
        return []

    selected = []
    invalid = []

    for choice in re.split(r"[,\s]+", raw_choices):
        if not choice:
            continue

        language = SEARCH_LANGUAGE_OPTIONS.get(choice)
        if language and language not in selected:
            selected.append(language)
        elif not language:
            invalid.append(choice)

    if invalid:
        print(f"Ignoring unrecognized choices: {', '.join(invalid)}")

    return selected


def translate_search_term(
    client: OpenAI,
    term: str,
    target_language: str,
    model: str,
    max_retries: int = 3,
) -> str:
    """Translate one query while preserving search syntax where practical."""

    prompt = f"""
Translate this social-media search term into {target_language}.

Return only the translated search term, with no quotation marks, label, or explanation.
Preserve hashtags, @handles, URLs, Boolean operators, and date/filter syntax exactly.
If the term is already in {target_language}, return it unchanged.

Search term:
{term}
""".strip()

    for attempt in range(1, max_retries + 1):
        try:
            translated = normalize_whitespace(
                generate_llm_text(client=client, model=model, prompt=prompt)
            )
            return translated.strip('"').strip("'")
        except Exception as e:
            if attempt == max_retries:
                print(
                    f"Could not translate {term!r} into {target_language}: {e}. "
                    "Skipping that query variant."
                )
                return ""

            time.sleep(2 * attempt)

    return ""


def build_interactive_search_terms() -> List[str]:
    """Prompt for terms/languages and return unique original + translated queries."""

    original_terms = prompt_for_search_terms(DEFAULT_SEARCH_TERMS)
    target_languages = prompt_for_search_languages()
    all_terms = list(original_terms)

    if target_languages:
        ensure_llm_configured()
        client = create_llm_client()
        print("\nTranslating search terms...")

        for term in original_terms:
            for language in target_languages:
                translated = translate_search_term(
                    client=client,
                    term=term,
                    target_language=language,
                    model=LLM_MODEL,
                )

                if translated and translated.casefold() not in {
                    existing.casefold() for existing in all_terms
                }:
                    all_terms.append(translated)
                    print(f"  {language}: {term!r} -> {translated!r}")

    print("\nSearch queries that will be used:")
    for term in all_terms:
        print(f"  - {term}")

    return all_terms


# ============================================================
# HTTP / NITTER FETCHING
# ============================================================

def create_requests_session() -> requests.Session:
    session = requests.Session()

    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 "
            "(Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 "
            "(KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive",
    })

    return session


def fetch_html(
    session: requests.Session,
    url: str,
    timeout: int = 30,
    max_retries: int = 2,
) -> Optional[str]:
    for attempt in range(1, max_retries + 1):
        try:
            print(f"Fetching page: {url}")

            response = session.get(
                url,
                timeout=timeout,
                verify=certifi.where(),
            )

            status = response.status_code
            text = response.text or ""

            print(f"Final URL: {response.url}")
            print(f"Status code: {status}")
            print(f"Response length: {len(text)} characters")
            print(f"Content-Type: {response.headers.get('content-type', '')}")

            if status != 200:
                preview = normalize_whitespace(text[:3000])
                print("\n--- NON-200 RESPONSE PREVIEW ---")
                print(preview)
                print("--- END PREVIEW ---\n")

            if status == 200:
                if len(text.strip()) == 0:
                    print("Instance returned a blank page.")
                    return None

                return text

            if status in [403, 404, 429, 503]:
                print(f"Instance returned {status}. Not bypassing; stopping this request.")
                return None

            print(f"Unexpected status code {status}. Attempt {attempt}/{max_retries}.")

        except requests.exceptions.RequestException as e:
            print(f"Request error on attempt {attempt}/{max_retries}: {repr(e)}")

        if attempt < max_retries:
            sleep_for = REQUEST_DELAY_SECONDS * attempt
            print(f"Sleeping {sleep_for} seconds before retry...")
            time.sleep(sleep_for)

    return None


def debug_nitter_page(html: str, max_chars: int = 3000):
    soup = BeautifulSoup(html, "html.parser")

    title = soup.find("title")
    title_text = title.get_text(" ", strip=True) if title else "[no title]"

    body_text = soup.get_text(" ", strip=True)
    body_text = normalize_whitespace(body_text)

    print("\n--- NITTER PAGE DEBUG ---")
    print("Title:", title_text)
    print("First body text:")
    print(body_text[:max_chars])
    print("\nDetected timeline item count:", len(soup.select(".timeline-item")))
    print("Detected tweet-content count:", len(soup.select(".tweet-content")))
    print("Detected timeline count:", len(soup.select(".timeline")))
    print("Detected show-more count:", len(soup.select(".show-more")))
    print("--- END DEBUG ---\n")


def find_next_nitter_page_url(
    html: str,
    current_url: str,
) -> Optional[str]:
    soup = BeautifulSoup(html, "html.parser")

    candidate_links = []

    candidate_links.extend(soup.select("div.show-more a"))
    candidate_links.extend(soup.select("a[href*='cursor=']"))

    for link in candidate_links:
        href = link.get("href")

        if not href:
            continue

        if "cursor=" in href or "search?" in href:
            return urljoin(current_url, href)

    return None


def parse_tweets_from_nitter_html(
    html: str,
    page_url: str,
    query: str,
) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    items = soup.select(".timeline-item")

    parsed = []

    for item in items:
        content_el = item.select_one(".tweet-content")

        if content_el is None:
            continue

        text = content_el.get_text("\n", strip=True)
        text = text.strip()

        if not text:
            continue

        username_el = item.select_one(".username")
        fullname_el = item.select_one(".fullname")
        date_link = item.select_one(".tweet-date a")
        tweet_link = item.select_one("a.tweet-link")

        username = get_text_or_empty(username_el).replace("@", "").strip()
        display_name = get_text_or_empty(fullname_el)

        date_raw = ""
        if date_link is not None:
            date_raw = date_link.get("title") or get_text_or_empty(date_link)

        date_iso = parse_nitter_date(date_raw)

        tweet_href = ""
        if tweet_link is not None:
            tweet_href = tweet_link.get("href") or ""

        nitter_url = urljoin(page_url, tweet_href) if tweet_href else ""
        x_url = make_x_url_from_nitter_url(nitter_url) if nitter_url else ""
        tweet_id = extract_tweet_id_from_url(nitter_url)

        stats_el = item.select_one(".tweet-stats")
        raw_stats = normalize_whitespace(get_text_or_empty(stats_el))

        retweet_header = item.select_one(".retweet-header")
        is_retweet = retweet_header is not None

        parsed.append({
            "query": query,
            "scraped_from": page_url,
            "nitter_url": nitter_url,
            "x_url": x_url,
            "tweet_id": tweet_id,
            "date_raw": date_raw,
            "date_iso": date_iso,
            "username": username,
            "display_name": display_name,
            "original_text": text,
            "raw_stats": raw_stats,
            "is_retweet": is_retweet,
        })

    return parsed


# ============================================================
# LLM TRANSLATION
# ============================================================

def translate_with_openai(
    client: OpenAI,
    text: str,
    source_lang: str,
    model: str = "gpt-5.6-luna",
    target_language: str = "English",
    max_retries: int = 3,
) -> str:
    if not isinstance(text, str) or not text.strip():
        return ""

    if source_lang == "en" and target_language.lower() == "english":
        return text

    prompt = f"""
Translate the following public social media post into {target_language}.

Rules:
- Preserve proper names, institution names, places, hashtags, @handles, emojis, and URLs.
- Preserve dates and numbers exactly.
- Do not summarize.
- Do not add commentary.
- If the post is already in {target_language}, return it unchanged.
- If the post contains mixed languages, translate only the non-{target_language} parts.
- Keep the tone close to the original.

Detected source language: {source_lang}

Post:
{text}
""".strip()

    for attempt in range(1, max_retries + 1):
        try:
            return generate_llm_text(client=client, model=model, prompt=prompt)

        except Exception as e:
            if attempt == max_retries:
                return f"[TRANSLATION ERROR: {e}] {text}"

            wait_time = 2 * attempt
            print(f"LLM translation error. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)

    return text


# ============================================================
# LOCATION INFERENCE + GEOCODING + MAP HELPERS
# ============================================================

def load_json_cache(path: str) -> dict:
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    return {}


def save_json_cache(path: str, data: dict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_language_geo_hint(lang_code: str) -> str:
    """
    Give the model a broad geographic prior based on detected language.
    These are intentionally broad and should not be treated as proof.
    """

    lang_code = (lang_code or "").lower().strip()

    hints = {
        "en": "English is global and should be treated as a weak location signal only.",
        "es": "Spanish suggests Spain or Latin America; use other evidence to distinguish.",
        "fr": "French suggests France, Belgium, Switzerland, Canada/Quebec, parts of West Africa, North Africa, or other Francophone regions.",
        "pt": "Portuguese suggests Portugal, Brazil, Angola, Mozambique, or other Lusophone regions.",
        "it": "Italian suggests Italy or Italian-language institutional contexts.",
        "de": "German suggests Germany, Austria, Switzerland, or German-language institutional contexts.",
        "ru": "Russian suggests Russia or Russian-speaking post-Soviet regions.",
        "ar": "Arabic suggests Arabic-speaking countries in the Middle East or North Africa.",
        "tr": "Turkish suggests Turkey or Turkish-speaking communities.",
        "ja": "Japanese suggests Japan or Japanese-language institutional contexts.",
        "ko": "Korean suggests South Korea or Korean-language institutional contexts.",
        "zh-cn": "Simplified Chinese suggests Mainland China, Singapore, Malaysia, or simplified-Chinese institutional contexts.",
        "zh-tw": "Traditional Chinese suggests Taiwan, Hong Kong, Macau, or traditional-Chinese institutional contexts.",
        "zh": "Chinese suggests a Chinese-language context; use script, institution names, and other evidence to distinguish Mainland China, Taiwan, Hong Kong, Macau, Singapore, Malaysia, or diaspora contexts.",
    }

    return hints.get(lang_code, "No strong language-based geographic prior available.")


def infer_location_with_openai(
    client: OpenAI,
    post: Dict[str, Any],
    model: str = "gpt-5.6-luna",
    max_retries: int = 3,
) -> Dict[str, Any]:
    """
    Infer a broad public location from tweet/account metadata.

    This intentionally avoids exact street-level locations.
    It should return city/region/country-level places only.
    """

    username = post.get("username", "")
    display_name = post.get("display_name", "")
    query = post.get("query", "")
    original_text = post.get("original_text", "")
    translated_text = post.get("translated_en", "")
    detected_language = post.get("detected_language", "unknown")
    language_geo_hint = get_language_geo_hint(detected_language)

    prompt = f"""
You are helping map public institutional social media posts.

Infer the most likely broad public location associated with this post.
Use only city, region, university, or country-level locations.
Do NOT infer private addresses, homes, street addresses, or precise personal locations.

Return ONLY valid JSON with these keys:
- location_name: string, a geocodable place such as "Madrid, Spain" or "University of Nairobi, Kenya"; use "" if unknown
- confidence: number between 0 and 1
- source: string, one of ["language", "username", "display_name", "tweet_text", "query", "combined", "unknown"]
- reason: short string explaining the evidence

Location inference rules:
1. Treat the language of the original post as a major geographic prior.
2. Use the detected language and script to narrow the likely region before considering weaker clues.
3. If the original language strongly aligns with a country, region, institution name, or search query, increase confidence.
4. If the language is widely used across many countries, such as English, French, Spanish, Portuguese, Arabic, or Chinese, do NOT choose a specific country based on language alone.
5. Prefer explicit place names in the original post text when present.
6. Next prefer institution names, university names, account names, or handles.
7. Use the translated text only to understand meaning; prioritize the original post for language, script, and named entities.
8. If language is the only evidence, return a broad country/region-level location with low confidence, usually 0.15 to 0.35.
9. If language plus account/institution/text evidence all point to the same place, confidence may be higher.
10. If no responsible location can be inferred, return location_name "" and confidence 0.

Detected language:
{detected_language}

Language-based geographic prior:
{language_geo_hint}

Metadata:
username: {username}
display_name: {display_name}
search_query: {query}

Original post:
{original_text}

Translated post:
{translated_text}
""".strip()

    for attempt in range(1, max_retries + 1):
        try:
            text = generate_llm_text(client=client, model=model, prompt=prompt)
            text = text.replace("```json", "").replace("```", "").strip()

            data = json.loads(text)

            return {
                "location_name": str(data.get("location_name", "")).strip(),
                "confidence": float(data.get("confidence", 0) or 0),
                "source": str(data.get("source", "unknown")).strip(),
                "reason": str(data.get("reason", "")).strip(),
            }

        except Exception as e:
            if attempt == max_retries:
                return {
                    "location_name": "",
                    "confidence": 0.0,
                    "source": "unknown",
                    "reason": f"Location inference failed: {e}",
                }

            time.sleep(2 * attempt)

    return {
        "location_name": "",
        "confidence": 0.0,
        "source": "unknown",
        "reason": "Location inference failed.",
    }


def geocode_location(
    location_name: str,
    cache: dict,
    geolocator: Nominatim,
    delay_seconds: float = 1.2,
) -> Dict[str, str]:
    """
    Geocode a location with Nominatim using local caching.
    """

    location_name = normalize_whitespace(location_name)

    if not location_name:
        return {
            "latitude": "",
            "longitude": "",
            "geocode_display_name": "",
        }

    if location_name in cache:
        return cache[location_name]

    try:
        print(f"Geocoding: {location_name}")

        result = geolocator.geocode(
            location_name,
            addressdetails=False,
            exactly_one=True,
            timeout=20,
        )

        time.sleep(delay_seconds)

        if result:
            data = {
                "latitude": str(result.latitude),
                "longitude": str(result.longitude),
                "geocode_display_name": result.address or location_name,
            }
        else:
            data = {
                "latitude": "",
                "longitude": "",
                "geocode_display_name": "",
            }

        cache[location_name] = data
        save_json_cache(GEOCODE_CACHE_FILE, cache)

        return data

    except (GeocoderTimedOut, GeocoderUnavailable) as e:
        print(f"Geocoding failed for {location_name}: {e}")

        return {
            "latitude": "",
            "longitude": "",
            "geocode_display_name": "",
        }

    except Exception as e:
        print(f"Unexpected geocoding error for {location_name}: {e}")

        return {
            "latitude": "",
            "longitude": "",
            "geocode_display_name": "",
        }


def enrich_records_with_locations(
    records: List[PostRecord],
    openai_model: str = "gpt-5.6-luna",
) -> List[PostRecord]:
    """
    Infer broad locations and geocode them.
    """

    if not records:
        return records

    if not INFER_LOCATIONS:
        return records

    print("\nInferring and geocoding locations...")

    client = create_llm_client()
    geolocator = Nominatim(user_agent=NOMINATIM_USER_AGENT)
    geocode_cache = load_json_cache(GEOCODE_CACHE_FILE)

    enriched_records = []

    iterator = records
    if tqdm is not None:
        iterator = tqdm(records)

    for record in iterator:
        post_dict = asdict(record)

        inference = infer_location_with_openai(
            client=client,
            post=post_dict,
            model=openai_model,
        )

        inferred_location = inference.get("location_name", "")
        confidence = float(inference.get("confidence", 0) or 0)
        source = inference.get("source", "unknown")
        reason = inference.get("reason", "")

        latitude = ""
        longitude = ""
        geocode_display_name = ""

        if inferred_location and confidence >= MIN_LOCATION_CONFIDENCE:
            geocode = geocode_location(
                location_name=inferred_location,
                cache=geocode_cache,
                geolocator=geolocator,
                delay_seconds=GEOCODE_DELAY_SECONDS,
            )

            latitude = geocode.get("latitude", "")
            longitude = geocode.get("longitude", "")
            geocode_display_name = geocode.get("geocode_display_name", "")

        record.inferred_location = inferred_location
        record.location_confidence = confidence
        record.location_source = source
        record.location_reason = reason
        record.latitude = latitude
        record.longitude = longitude
        record.geocode_display_name = geocode_display_name

        enriched_records.append(record)

    return enriched_records


def create_tweet_map(
    df: pd.DataFrame,
    map_file: str = f"social_search_map_{RUN_TIMESTAMP}.html",
):
    """
    Create an interactive Folium map with popups showing original and translated tweets.

    Tile fix:
    - Do not use Folium's direct OpenStreetMap default tile layer.
    - That default can trigger 503r / Access Blocked errors from the OSM tile server.
    - Use CARTO basemap tiles instead, with proper attribution.
    """

    if df is None or df.empty:
        print("No dataframe available for mapping.")
        return None

    if "latitude" not in df.columns or "longitude" not in df.columns:
        print("No latitude/longitude columns found. Skipping map.")
        return None

    map_df = df.copy()

    map_df["latitude_num"] = pd.to_numeric(map_df["latitude"], errors="coerce")
    map_df["longitude_num"] = pd.to_numeric(map_df["longitude"], errors="coerce")

    map_df = map_df.dropna(subset=["latitude_num", "longitude_num"])

    if map_df.empty:
        print("No geocoded rows available for mapping.")
        return None

    center_lat = map_df["latitude_num"].mean()
    center_lon = map_df["longitude_num"].mean()

    # IMPORTANT:
    # tiles=None prevents Folium from adding its default OpenStreetMap layer.
    # The default OSM tile endpoint is the source of the 503r access-blocked errors.
    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=2,
        tiles=None,
        control_scale=True,
    )

    # Primary basemap: CARTO Positron. This avoids direct requests to
    # tile.openstreetmap.org while still providing a clean global basemap.
    folium.TileLayer(
        tiles=MAP_BASE_TILE_URL,
        attr=MAP_BASE_TILE_ATTRIBUTION,
        name=MAP_BASE_TILE_NAME,
        overlay=False,
        control=True,
    ).add_to(m)

    # Optional alternate basemaps. These are useful if one tile provider is slow
    # or temporarily unavailable. None of these uses the direct OSM tile server.
    if ADD_ALTERNATE_BASEMAPS:
        folium.TileLayer(
            tiles="https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png",
            attr=MAP_BASE_TILE_ATTRIBUTION,
            name="CartoDB Dark Matter",
            overlay=False,
            control=True,
        ).add_to(m)

        folium.TileLayer(
            tiles=(
                "https://server.arcgisonline.com/ArcGIS/rest/services/"
                "World_Imagery/MapServer/tile/{z}/{y}/{x}"
            ),
            attr=(
                "Tiles &copy; Esri &mdash; Source: Esri, i-cubed, USDA, USGS, "
                "AEX, GeoEye, Getmapping, Aerogrid, IGN, IGP, UPR-EGP, and the GIS User Community"
            ),
            name="Esri World Imagery",
            overlay=False,
            control=True,
        ).add_to(m)

    cluster = MarkerCluster(name="Inferred tweet locations").add_to(m)

    for _, row in map_df.iterrows():
        original_text = html_lib.escape(str(row.get("original_text", "")))
        translated_en = html_lib.escape(str(row.get("translated_en", "")))
        username = html_lib.escape(str(row.get("username", "")))
        display_name = html_lib.escape(str(row.get("display_name", "")))
        date_raw = html_lib.escape(str(row.get("date_raw", "")))
        inferred_location = html_lib.escape(str(row.get("inferred_location", "")))
        confidence = html_lib.escape(str(row.get("location_confidence", "")))
        source = html_lib.escape(str(row.get("location_source", "")))
        reason = html_lib.escape(str(row.get("location_reason", "")))
        post_url = html_lib.escape(
            str(row.get("post_url", row.get("x_url", "")))
        )
        source_mode = str(row.get("source_mode", ""))
        source_url = html_lib.escape(str(row.get("source_url", "")))
        legacy_link = (
            f'<div><a href="{source_url}" target="_blank">Open saved Nitter source</a></div>'
            if source_mode == "saved_html" and source_url
            else ""
        )

        popup_html = f"""
        <div style="width: 380px; font-family: Arial, sans-serif;">
            <h4 style="margin-bottom: 4px;">@{username}</h4>
            <div><strong>Name:</strong> {display_name}</div>
            <div><strong>Date:</strong> {date_raw}</div>
            <div><strong>Inferred location:</strong> {inferred_location}</div>
            <div><strong>Confidence:</strong> {confidence}</div>
            <div><strong>Source:</strong> {source}</div>
            <div><strong>Reason:</strong> {reason}</div>
            <hr>
            <div><strong>Original:</strong></div>
            <div style="white-space: pre-wrap; margin-bottom: 8px;">{original_text}</div>
            <div><strong>Translation:</strong></div>
            <div style="white-space: pre-wrap; margin-bottom: 8px;">{translated_en}</div>
            <hr>
            <div><a href="{post_url}" target="_blank">Open original post</a></div>
            {legacy_link}
        </div>
        """

        tooltip = f"@{username} — {inferred_location}"

        folium.Marker(
            location=[row["latitude_num"], row["longitude_num"]],
            popup=folium.Popup(popup_html, max_width=440),
            tooltip=tooltip,
        ).add_to(cluster)

    folium.LayerControl().add_to(m)

    m.save(map_file)

    print(f"Map saved to: {map_file}")
    print("Map tile layer: CARTO basemap CDN, not the direct OpenStreetMap tile server.")

    return m


# ============================================================
# RECORD CREATION
# ============================================================

def make_record_from_post(
    post: Dict[str, Any],
    source_mode: str,
    source_host: str,
    client: Optional[OpenAI],
    openai_model: str,
    target_language: str,
    translate: bool,
) -> PostRecord:
    original_text = post.get("original_text", "")
    detected_language = detect_language_safe(original_text)

    if translate and client is not None:
        translated_en = translate_with_openai(
            client=client,
            text=original_text,
            source_lang=detected_language,
            model=openai_model,
            target_language=target_language,
        )
        time.sleep(TRANSLATION_DELAY_SECONDS)
    else:
        translated_en = ""

    return PostRecord(
        platform=post.get("platform", "legacy"),
        query=post.get("query", ""),
        source_mode=source_mode,
        source_host=source_host,
        source_url=post.get("source_url", post.get("nitter_url", "")),
        scraped_from=post.get("scraped_from", ""),
        post_url=post.get("post_url", post.get("x_url", "")),
        x_url=post.get("x_url", ""),
        tweet_id=post.get("tweet_id", ""),
        date_raw=post.get("date_raw", ""),
        date_iso=post.get("date_iso", ""),
        username=post.get("username", ""),
        display_name=post.get("display_name", ""),
        detected_language=detected_language,
        original_text=original_text,
        translated_en=translated_en,
        raw_stats=post.get("raw_stats", ""),
        is_retweet=post.get("is_retweet", False),
        inferred_location="",
        location_confidence=0.0,
        location_source="",
        location_reason="",
        latitude="",
        longitude="",
        geocode_display_name="",
    )


# ============================================================
# X API COLLECTION
# ============================================================

def run_x_api_collection(
    bearer_token: str,
    since: Optional[str] = None,
    until: Optional[str] = None,
    max_posts_per_query: int = 5,
    max_pages_per_query: int = 1,
    output_file: str = f"social_search_posts_{RUN_TIMESTAMP}.csv",
    handles: Optional[List[str]] = None,
    search_terms: Optional[List[str]] = None,
    include_retweets: bool = False,
    target_language: str = "English",
    openai_model: str = "gpt-5.6-luna",
    translate: bool = False,
):
    print("Official X API collection started.")

    session = create_requests_session()

    if translate:
        print("Translation is ON.")
        client = create_llm_client()
    else:
        print("Translation is OFF.")
        client = None

    if search_terms is None:
        search_terms = DEFAULT_SEARCH_TERMS

    queries = build_queries(
        terms=search_terms,
        handles=handles,
    )

    print(f"\nBuilt {len(queries)} queries:")
    for q in queries:
        print("  ", q)

    records: List[PostRecord] = []
    seen_ids = set()
    seen_text_fallbacks = set()

    for query in queries:
        print("\n" + "=" * 70)
        api_query = build_x_query(query, include_retweets=include_retweets)
        print(f"Searching X for: {api_query}")

        posts_collected_for_query = 0
        next_token = None

        for page_num in range(1, max_pages_per_query + 1):
            print(f"Page {page_num}/{max_pages_per_query}")
            remaining = max_posts_per_query - posts_collected_for_query

            try:
                parsed_posts, next_token, request_url = fetch_x_api_page(
                    session=session,
                    bearer_token=bearer_token,
                    query=api_query,
                    max_results=remaining,
                    since=since,
                    until=until,
                    next_token=next_token,
                )
            except RuntimeError as exc:
                print(f"X API error: {exc}")
                break

            print(f"X returned {len(parsed_posts)} posts.")

            for post in parsed_posts:
                tweet_id = post.get("tweet_id", "")
                fallback_key = (
                    post.get("username", ""),
                    post.get("date_raw", ""),
                    post.get("original_text", "")[:120],
                )

                if tweet_id and tweet_id in seen_ids:
                    continue
                if not tweet_id and fallback_key in seen_text_fallbacks:
                    continue

                if tweet_id:
                    seen_ids.add(tweet_id)
                else:
                    seen_text_fallbacks.add(fallback_key)

                record = make_record_from_post(
                    post=post,
                    source_mode="x_api",
                    source_host="api.x.com",
                    client=client,
                    openai_model=openai_model,
                    target_language=target_language,
                    translate=translate,
                )
                records.append(record)
                posts_collected_for_query += 1

                print(
                    f"Collected {posts_collected_for_query}/{max_posts_per_query} "
                    f"for this query: @{record.username} {record.date_raw}"
                )
                if posts_collected_for_query >= max_posts_per_query:
                    break

            if posts_collected_for_query >= max_posts_per_query:
                print("Reached max posts for this query.")
                break
            if not next_token:
                print("No additional result page is available.")
                break

            print(f"Sleeping {REQUEST_DELAY_SECONDS} seconds before next page...")
            time.sleep(REQUEST_DELAY_SECONDS)

    return records


def run_bluesky_collection(
    since: Optional[str] = None,
    until: Optional[str] = None,
    max_posts_per_query: int = 5,
    max_pages_per_query: int = 1,
    search_terms: Optional[List[str]] = None,
    target_language: str = "English",
    openai_model: str = "gpt-5.6-luna",
    translate: bool = False,
    access_jwt: str = "",
) -> List[PostRecord]:
    """Collect public Bluesky posts and normalize them into PostRecord objects."""
    mode_label = "authenticated PDS proxy" if access_jwt else "public API"
    print(f"Bluesky {mode_label} collection started.")
    session = create_requests_session()
    client = create_llm_client() if translate else None
    queries = build_queries(terms=search_terms or DEFAULT_SEARCH_TERMS)
    records: List[PostRecord] = []
    seen_ids = set()

    for query in queries:
        print("\n" + "=" * 70)
        print(f"Searching Bluesky for: {query}")
        collected = 0
        cursor = None

        for page_num in range(1, max_pages_per_query + 1):
            try:
                posts, cursor, _ = fetch_bluesky_page(
                    session=session,
                    query=query,
                    max_results=max_posts_per_query - collected,
                    since=since,
                    until=until,
                    cursor=cursor,
                    access_jwt=access_jwt,
                )
            except RuntimeError as exc:
                print(f"Bluesky API error: {exc}")
                break

            print(f"Bluesky returned {len(posts)} posts on page {page_num}.")
            for post in posts:
                post_id = post.get("tweet_id", "")
                if post_id and ("bluesky", post_id) in seen_ids:
                    continue
                if post_id:
                    seen_ids.add(("bluesky", post_id))

                record = make_record_from_post(
                    post=post,
                    source_mode="api",
                    source_host=("bsky.social" if access_jwt else "public.api.bsky.app"),
                    client=client,
                    openai_model=openai_model,
                    target_language=target_language,
                    translate=translate,
                )
                records.append(record)
                collected += 1
                print(
                    f"Collected {collected}/{max_posts_per_query} from Bluesky: "
                    f"@{record.username} {record.date_raw}"
                )
                if collected >= max_posts_per_query:
                    break

            if collected >= max_posts_per_query or not cursor:
                break
            time.sleep(REQUEST_DELAY_SECONDS)

    return records


def run_mastodon_collection(
    instance_url: str,
    access_token: str,
    since: Optional[str] = None,
    until: Optional[str] = None,
    max_posts_per_query: int = 5,
    max_pages_per_query: int = 1,
    search_terms: Optional[List[str]] = None,
    include_reblogs: bool = False,
    target_language: str = "English",
    openai_model: str = "gpt-5.6-luna",
    translate: bool = False,
) -> List[PostRecord]:
    """Collect searchable public statuses known to one Mastodon instance."""
    print(f"Mastodon API collection started for {instance_url}.")
    session = create_requests_session()
    client = create_llm_client() if translate else None
    queries = build_queries(terms=search_terms or DEFAULT_SEARCH_TERMS)
    records: List[PostRecord] = []
    seen_ids = set()

    for query in queries:
        print("\n" + "=" * 70)
        print(f"Searching Mastodon on {instance_url} for: {query}")
        collected = 0
        offset = 0

        for page_num in range(1, max_pages_per_query + 1):
            try:
                posts, next_offset, _ = fetch_mastodon_page(
                    session=session,
                    instance_url=instance_url,
                    access_token=access_token,
                    query=query,
                    max_results=max_posts_per_query - collected,
                    since=since,
                    until=until,
                    offset=offset,
                )
            except RuntimeError as exc:
                print(f"Mastodon API error: {exc}")
                break

            print(f"Mastodon returned {len(posts)} usable posts on page {page_num}.")
            for post in posts:
                if not include_reblogs and post.get("is_retweet"):
                    continue
                post_id = post.get("tweet_id", "")
                dedupe_key = (instance_url, post_id)
                if post_id and dedupe_key in seen_ids:
                    continue
                if post_id:
                    seen_ids.add(dedupe_key)

                record = make_record_from_post(
                    post=post,
                    source_mode="api",
                    source_host=urlparse(instance_url).netloc,
                    client=client,
                    openai_model=openai_model,
                    target_language=target_language,
                    translate=translate,
                )
                records.append(record)
                collected += 1
                print(
                    f"Collected {collected}/{max_posts_per_query} from Mastodon: "
                    f"@{record.username} {record.date_raw}"
                )
                if collected >= max_posts_per_query:
                    break

            if collected >= max_posts_per_query or next_offset is None:
                break
            offset = next_offset
            time.sleep(REQUEST_DELAY_SECONDS)

    return records


# ============================================================
# SAVED HTML COLLECTION
# ============================================================

def run_saved_html_collection(
    html_files: List[str],
    output_file: str,
    include_retweets: bool = False,
    target_language: str = "English",
    openai_model: str = "gpt-5.6-luna",
    translate: bool = False,
):
    print("Saved HTML collection started.")

    if translate:
        print("Translation is ON.")
        client = create_llm_client()
    else:
        print("Translation is OFF.")
        client = None

    records: List[PostRecord] = []
    seen_ids = set()
    seen_text_fallbacks = set()

    for html_file in html_files:
        print("\n" + "=" * 70)
        print(f"Parsing saved HTML file: {html_file}")

        if not os.path.exists(html_file):
            print(f"File not found: {html_file}")
            continue

        with open(html_file, "r", encoding="utf-8", errors="replace") as f:
            html = f.read()

        debug_nitter_page(html, max_chars=1500)

        parsed_posts = parse_tweets_from_nitter_html(
            html=html,
            page_url="https://nitter.net/search",
            query=f"saved_html:{html_file}",
        )

        print(f"Parsed {len(parsed_posts)} posts from saved HTML.")

        for post in parsed_posts:
            if not include_retweets and post.get("is_retweet"):
                print("Skipping retweet.")
                continue

            tweet_id = post.get("tweet_id", "")

            fallback_key = (
                post.get("username", ""),
                post.get("date_raw", ""),
                post.get("original_text", "")[:120],
            )

            if tweet_id and tweet_id in seen_ids:
                continue

            if not tweet_id and fallback_key in seen_text_fallbacks:
                continue

            if tweet_id:
                seen_ids.add(tweet_id)
            else:
                seen_text_fallbacks.add(fallback_key)

            record = make_record_from_post(
                post=post,
                source_mode="saved_html",
                source_host="saved_html",
                client=client,
                openai_model=openai_model,
                target_language=target_language,
                translate=translate,
            )

            records.append(record)

            print(f"Collected saved HTML post: @{record.username} {record.date_raw}")

    return records


# ============================================================
# SAVE OUTPUT
# ============================================================

def save_records(records: List[PostRecord], output_file: str):
    if not records:
        print("\nNo records collected.")
        print("Likely causes:")
        print("1. The selected APIs returned no matching accessible posts.")
        print("2. The query is too narrow.")
        print("3. Date filters exclude available results.")
        print("4. A credential, API access, billing, or rate-limit error occurred.")
        print("5. Mastodon full-text search is unavailable on the chosen server.")
        print("6. In legacy mode, saved HTML contains no visible timeline items.")
        return None

    if CREATE_MAP or INFER_LOCATIONS:
        records = enrich_records_with_locations(
            records=records,
            openai_model=LLM_MODEL,
        )

    df = pd.DataFrame([asdict(r) for r in records])

    # Clean all cells for CSV and Excel readability
    for col in df.columns:
        df[col] = df[col].apply(clean_cell)

    # Put the most useful columns first
    preferred_order = [
        "platform",
        "date_iso",
        "date_raw",
        "username",
        "display_name",
        "detected_language",
        "inferred_location",
        "location_confidence",
        "location_source",
        "location_reason",
        "latitude",
        "longitude",
        "geocode_display_name",
        "original_text",
        "translated_en",
        "post_url",
        "x_url",
        "tweet_id",
        "raw_stats",
        "is_retweet",
        "query",
        "source_mode",
        "source_host",
        "source_url",
        "scraped_from",
    ]

    existing_order = [col for col in preferred_order if col in df.columns]
    remaining_cols = [col for col in df.columns if col not in existing_order]
    df = df[existing_order + remaining_cols]

    # Make filenames
    if output_file.lower().endswith(".csv"):
        csv_file = output_file
        xlsx_file = output_file[:-4] + ".xlsx"
    else:
        csv_file = output_file + ".csv"
        xlsx_file = output_file + ".xlsx"

    # Clean CSV export
    df.to_csv(
        csv_file,
        index=False,
        encoding="utf-8-sig",
        quoting=csv.QUOTE_ALL,
        lineterminator="\n",
    )

    # Formatted Excel export
    with pd.ExcelWriter(xlsx_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="posts")

        worksheet = writer.sheets["posts"]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        # Header styling
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Set column widths by column name
        width_by_column_name = {
            "platform": 14,
            "date_iso": 24,
            "date_raw": 28,
            "username": 20,
            "display_name": 30,
            "detected_language": 18,
            "inferred_location": 34,
            "location_confidence": 18,
            "location_source": 18,
            "location_reason": 60,
            "latitude": 18,
            "longitude": 18,
            "geocode_display_name": 60,
            "original_text": 90,
            "translated_en": 90,
            "post_url": 60,
            "x_url": 60,
            "tweet_id": 24,
            "raw_stats": 30,
            "is_retweet": 14,
            "query": 30,
            "source_mode": 16,
            "source_host": 24,
            "source_url": 60,
            "scraped_from": 60,
        }

        for idx, column_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = width_by_column_name.get(
                column_name,
                24,
            )

        # Wrap and top-align all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # Make rows taller for readability
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 60

    print(f"\nDone. Wrote {len(df)} records.")
    print(f"CSV file:   {csv_file}")
    print(f"Excel file: {xlsx_file}")

    if CREATE_MAP:
        create_tweet_map(
            df=df,
            map_file=MAP_FILE,
        )

    print("\nPreview:")
    print(df.head())

    return df


# ============================================================
# RUN THE SCRIPT
# ============================================================

if MODE in {"api", "x_api"}:
    selected_sources = prompt_data_sources()
    if "x" in selected_sources:
        X_BEARER_TOKEN = prompt_x_bearer_token()
    if "bluesky" in selected_sources:
        BLUESKY_IDENTIFIER, BLUESKY_APP_PASSWORD = prompt_bluesky_settings()
        if BLUESKY_IDENTIFIER:
            print("Signing in to Bluesky with the app password...")
            try:
                BLUESKY_ACCESS_JWT = create_bluesky_access_token(
                    create_requests_session(),
                    BLUESKY_IDENTIFIER,
                    BLUESKY_APP_PASSWORD,
                )
                print("Bluesky authentication succeeded.")
            except RuntimeError as exc:
                print(f"Bluesky authentication error: {exc}")
                print("Continuing with public-only Bluesky search.")
    if "mastodon" in selected_sources:
        MASTODON_INSTANCE_URL, MASTODON_ACCESS_TOKEN = prompt_mastodon_settings()

    if TRANSLATE_POSTS or INFER_LOCATIONS:
        ensure_llm_configured()
    else:
        print(
            "Post translation/location inference are off. An LLM key will be "
            "requested only if you choose translated search terms."
        )

    active_search_terms = build_interactive_search_terms()
    collected_records: List[PostRecord] = []

    if "x" in selected_sources:
        collected_records.extend(
            run_x_api_collection(
                bearer_token=X_BEARER_TOKEN,
                since=SINCE_DATE,
                until=UNTIL_DATE,
                max_posts_per_query=MAX_POSTS_PER_QUERY,
                max_pages_per_query=MAX_PAGES_PER_QUERY,
                output_file=OUTPUT_FILE,
                handles=SEARCH_HANDLES if SEARCH_HANDLES else None,
                search_terms=active_search_terms,
                include_retweets=INCLUDE_RETWEETS,
                target_language=TARGET_LANGUAGE,
                openai_model=LLM_MODEL,
                translate=TRANSLATE_POSTS,
            )
        )

    if "bluesky" in selected_sources:
        collected_records.extend(
            run_bluesky_collection(
                since=SINCE_DATE,
                until=UNTIL_DATE,
                max_posts_per_query=MAX_POSTS_PER_QUERY,
                max_pages_per_query=MAX_PAGES_PER_QUERY,
                search_terms=active_search_terms,
                target_language=TARGET_LANGUAGE,
                openai_model=LLM_MODEL,
                translate=TRANSLATE_POSTS,
                access_jwt=BLUESKY_ACCESS_JWT,
            )
        )

    if "mastodon" in selected_sources:
        collected_records.extend(
            run_mastodon_collection(
                instance_url=MASTODON_INSTANCE_URL,
                access_token=MASTODON_ACCESS_TOKEN,
                since=SINCE_DATE,
                until=UNTIL_DATE,
                max_posts_per_query=MAX_POSTS_PER_QUERY,
                max_pages_per_query=MAX_PAGES_PER_QUERY,
                search_terms=active_search_terms,
                include_reblogs=INCLUDE_RETWEETS,
                target_language=TARGET_LANGUAGE,
                openai_model=LLM_MODEL,
                translate=TRANSLATE_POSTS,
            )
        )

    df = save_records(collected_records, OUTPUT_FILE)

elif MODE == "saved_html":
    if TRANSLATE_POSTS or INFER_LOCATIONS:
        ensure_llm_configured()
    collected_records = run_saved_html_collection(
        html_files=SAVED_HTML_FILES,
        output_file=OUTPUT_FILE,
        include_retweets=INCLUDE_RETWEETS,
        target_language=TARGET_LANGUAGE,
        openai_model=LLM_MODEL,
        translate=TRANSLATE_POSTS,
    )
    df = save_records(collected_records, OUTPUT_FILE)

else:
    raise ValueError("MODE must be either 'api' or 'saved_html'.")
